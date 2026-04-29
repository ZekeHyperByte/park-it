"""Report export service — CSV, Excel, PDF."""

import csv
import io
from datetime import date

from api.app.schemas.report import SummaryReport


def export_summary_csv(report: SummaryReport) -> bytes:
    """Export summary report as CSV."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Metric", "Value"])
    writer.writerow(["Total Transactions", report.total_transactions])
    writer.writerow(["Total Revenue", report.total_revenue])
    writer.writerow(["Cash Revenue", report.cash_revenue])
    writer.writerow(["E-Money Revenue", report.emoney_revenue])
    writer.writerow(["RFID Revenue", report.rfid_revenue])
    writer.writerow(["Average Fee", report.average_fee])
    writer.writerow(["Active Transactions", report.active_transactions])
    writer.writerow(["Completed Transactions", report.completed_transactions])
    return output.getvalue().encode("utf-8-sig")


def export_summary_xlsx(report: SummaryReport) -> bytes:
    """Export summary report as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "E-Parking Summary Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    rows = [
        ["Metric", "Value"],
        ["Total Transactions", report.total_transactions],
        ["Total Revenue", report.total_revenue],
        ["Cash Revenue", report.cash_revenue],
        ["E-Money Revenue", report.emoney_revenue],
        ["RFID Revenue", report.rfid_revenue],
        ["Average Fee", report.average_fee],
        ["Active Transactions", report.active_transactions],
        ["Completed Transactions", report.completed_transactions],
    ]

    for i, row in enumerate(rows, start=3):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == 3:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 25

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_summary_pdf(
    report: SummaryReport,
    site_name: str = "E-Parking",
    date_from: date | None = None,
    date_to: date | None = None,
) -> bytes:
    """Export summary report as PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>{site_name}</b>", styles["Title"]))
    if date_from and date_to:
        elements.append(Paragraph(f"Period: {date_from} to {date_to}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    data = [
        ["Metric", "Value"],
        ["Total Transactions", f"{report.total_transactions:,}"],
        ["Total Revenue", f"Rp {report.total_revenue:,}"],
        ["Cash Revenue", f"Rp {report.cash_revenue:,}"],
        ["E-Money Revenue", f"Rp {report.emoney_revenue:,}"],
        ["RFID Revenue", f"Rp {report.rfid_revenue:,}"],
        ["Average Fee", f"Rp {report.average_fee:,.2f}"],
        ["Active Transactions", f"{report.active_transactions:,}"],
        ["Completed Transactions", f"{report.completed_transactions:,}"],
    ]

    table = Table(data, colWidths=[250, 200])
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ])
    )

    elements.append(table)
    doc.build(elements)
    return output.getvalue()
